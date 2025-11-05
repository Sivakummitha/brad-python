from openpyxl import Workbook, load_workbook
import re
wb = Workbook()
ws = wb.active
ws.title = "Numbers"
ws.append(["Decimal", "Binary", "Octal", "Hexadecimal"])
for n in range(1, 21):
    ws.append([n, bin(n)[2:], oct(n)[2:], hex(n)[2:].upper()])
wb.save("numbers.xlsx")
wb = load_workbook("numbers.xlsx")
ws = wb.active
def is_valid_binary(v): return bool(re.fullmatch(r'[01]+', v))
def is_valid_octal(v): return bool(re.fullmatch(r'[0-7]+', v))
def is_valid_hex(v): return bool(re.fullmatch(r'[0-9A-Fa-f]+', v))
for row in ws.iter_rows(min_row=2, values_only=True):
    d, b, o, h = row
    vb, vo, vh = is_valid_binary(b), is_valid_octal(o), is_valid_hex(h)
    cb = int(b, 2) == d if vb else False
    co = int(o, 8) == d if vo else False
    ch = int(h, 16) == d if vh else False
    print(f"Decimal: {d}")
    print(f" Binary({b}) - {'Valid' if vb else 'Invalid'}, {'Correct' if cb else 'Incorrect'}")
    print(f" Octal({o}) - {'Valid' if vo else 'Invalid'}, {'Correct' if co else 'Incorrect'}")
    print(f" Hex({h}) - {'Valid' if vh else 'Invalid'}, {'Correct' if ch else 'Incorrect'}\n")
