import openpyxl
from openpyxl import Workbook
wb=Workbook()
ws=wb.active
ws.title="mydata"
ws['A1']="sivacharan"
ws['A2']="mona"
ws['A3']="rahul"
ws['A4']="srinu"
ws['A5']="baahubali"
wb.save("sample_data.xlsx")
ws['A5']="ramya"
wb.save("sample_data.xlsx")
ws.merge_cells("A1:B1")
